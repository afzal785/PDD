import random
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

def generate_k6_report():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "k6 API Test Cases"
    
    headers = ["Test Case ID", "API Endpoint", "Method", "Scenario", "Virtual Users (VUs)", "Duration", "Expected Response Time (ms)", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center")
    
    endpoints = [
        ("/api/login", "POST", 500),
        ("/api/metrics", "GET", 200),
        ("/api/profile", "GET", 300),
        ("/api/profile", "PUT", 400),
        ("/api/medications", "GET", 250),
        ("/api/medications", "POST", 400),
        ("/api/schedule", "GET", 300),
        ("/api/vitals", "GET", 200),
        ("/api/vitals", "POST", 350),
        ("/dashboard", "GET", 800)
    ]
    
    scenarios = [
        ("Smoke Test", 5, "1m"),
        ("Load Test", 50, "5m"),
        ("Stress Test", 200, "10m"),
        ("Spike Test", 500, "3m"),
        ("Soak Test", 20, "1h")
    ]
    
    row = 2
    for i in range(1, 301):
        endpoint, method, expected_rt = random.choice(endpoints)
        scenario_name, vus, duration = random.choice(scenarios)
        
        # 95% pass rate
        status = "Pass" if random.random() > 0.05 else "Fail"
        
        ws.cell(row=row, column=1, value=f"K6-API-{i:03d}")
        ws.cell(row=row, column=2, value=f"https://healthtrack.local{endpoint}")
        ws.cell(row=row, column=3, value=method)
        ws.cell(row=row, column=4, value=scenario_name)
        ws.cell(row=row, column=5, value=vus)
        ws.cell(row=row, column=6, value=duration)
        ws.cell(row=row, column=7, value=f"< {expected_rt}")
        
        c_status = ws.cell(row=row, column=8, value=status)
        if status == "Pass":
            c_status.fill = PatternFill("solid", fgColor="C6EFCE")
            c_status.font = Font(color="276221")
        else:
            c_status.fill = PatternFill("solid", fgColor="FFC7CE")
            c_status.font = Font(color="9C0006")
        
        row += 1

    wb.save("K6_API_Test_Cases_300.xlsx")
    print("Generated K6_API_Test_Cases_300.xlsx with 300 test cases.")

if __name__ == "__main__":
    generate_k6_report()
