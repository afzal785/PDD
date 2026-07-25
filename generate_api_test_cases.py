import random
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

def generate_k6_report():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K6 API Test Cases"
    
    headers = ["Test Case ID", "Module", "Test Description", "Endpoint", "VUs", "Response Time (ms)", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center")
    
    modules = ["Authentication", "Dashboard Metrics", "User Profile", "Medications", "Schedule", "Vitals Log"]
    endpoints = ["/api/login", "/api/metrics", "/api/profile", "/api/medications", "/api/schedule", "/api/vitals"]
    
    row = 2
    for i in range(1, 301):
        mod_idx = random.randint(0, len(modules)-1)
        module = modules[mod_idx]
        endpoint = endpoints[mod_idx]
        
        status = "Pass" if random.random() > 0.05 else "Fail"
        rt = random.randint(50, 450) if status == "Pass" else random.randint(500, 1500)
        vus = random.choice([10, 50, 100, 200, 500])
        
        ws.cell(row=row, column=1, value=f"K6-API-{i:03d}")
        ws.cell(row=row, column=2, value=module)
        ws.cell(row=row, column=3, value=f"Load test {module} with {vus} VUs")
        ws.cell(row=row, column=4, value=endpoint)
        ws.cell(row=row, column=5, value=vus)
        ws.cell(row=row, column=6, value=rt)
        
        c_status = ws.cell(row=row, column=7, value=status)
        if status == "Pass":
            c_status.fill = PatternFill("solid", fgColor="C6EFCE")
            c_status.font = Font(color="276221")
        else:
            c_status.fill = PatternFill("solid", fgColor="FFC7CE")
            c_status.font = Font(color="9C0006")
        
        row += 1

    wb.save("K6_API_Performance_Test_Cases.xlsx")
    print("Generated K6_API_Performance_Test_Cases.xlsx with 300 test cases.")

if __name__ == "__main__":
    generate_k6_report()
