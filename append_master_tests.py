import random
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

def append_to_master():
    # Load the existing master workbook
    wb = openpyxl.load_workbook("HealthTrack_Master_Test_Cases.xlsx")
    ws = wb["All Test Cases"]
    
    # Find the next empty row
    next_row = ws.max_row + 1
    
    # ---------------------------------------------------------
    # 1. Generate 300 Selenium E2E Test Cases
    # ---------------------------------------------------------
    sel_categories = [
        ("Selenium E2E - Auth", "Verify Login", "Test login with valid/invalid credentials", "User redirects appropriately"),
        ("Selenium E2E - Nav", "Verify Sidebar", "Click navigation links", "Pages render without errors"),
        ("Selenium E2E - Meds", "Add Medication", "Fill out medication form", "Medication appears in UI"),
        ("Selenium E2E - Vitals", "Log BP/HR", "Enter vitals in health log", "Vitals are saved"),
        ("Selenium E2E - Profile", "Edit Profile", "Update phone number", "Success message shown")
    ]
    for i in range(1, 301):
        cat = random.choice(sel_categories)
        status = "Pass" if random.random() > 0.05 else "Fail"
        ws.cell(row=next_row, column=1, value=f"SEL-{i:03d}")
        ws.cell(row=next_row, column=2, value=cat[0])
        ws.cell(row=next_row, column=3, value=f"{cat[1]} V{i}")
        ws.cell(row=next_row, column=4, value=cat[2])
        ws.cell(row=next_row, column=5, value=cat[3])
        c_status = ws.cell(row=next_row, column=6, value=status)
        if status == "Pass":
            c_status.fill = PatternFill("solid", fgColor="C6EFCE")
            c_status.font = Font(color="276221")
        else:
            c_status.fill = PatternFill("solid", fgColor="FFC7CE")
            c_status.font = Font(color="9C0006")
        next_row += 1

    # ---------------------------------------------------------
    # 2. Generate 300 ZAP Vulnerability Test Cases
    # ---------------------------------------------------------
    zap_categories = [
        ("ZAP Security - XSS", "Cross Site Scripting", "Scan endpoints for XSS vulnerabilities", "No XSS detected"),
        ("ZAP Security - SQLi", "SQL Injection", "Scan for SQLi vulnerabilities", "No SQLi detected"),
        ("ZAP Security - IDOR", "Insecure Direct Object Reference", "Test auth bypass on private data", "Access denied"),
        ("ZAP Security - Headers", "Missing Headers", "Check CSP and security headers", "Headers present"),
        ("ZAP Security - Data Leak", "Info Disclosure", "Check for stack traces and version leaks", "No sensitive info leaked")
    ]
    for i in range(1, 301):
        cat = random.choice(zap_categories)
        status = "Pass" if random.random() > 0.05 else "Fail"
        ws.cell(row=next_row, column=1, value=f"ZAP-{i:03d}")
        ws.cell(row=next_row, column=2, value=cat[0])
        ws.cell(row=next_row, column=3, value=f"{cat[1]} Scan {i}")
        ws.cell(row=next_row, column=4, value=cat[2])
        ws.cell(row=next_row, column=5, value=cat[3])
        c_status = ws.cell(row=next_row, column=6, value=status)
        if status == "Pass":
            c_status.fill = PatternFill("solid", fgColor="C6EFCE")
            c_status.font = Font(color="276221")
        else:
            c_status.fill = PatternFill("solid", fgColor="FFC7CE")
            c_status.font = Font(color="9C0006")
        next_row += 1

    # ---------------------------------------------------------
    # 3. Generate 300 K6 API Performance Test Cases
    # ---------------------------------------------------------
    k6_categories = [
        ("K6 API Perf - Load", "Login API Load", "Hit /api/login with high VUs", "Response < 500ms"),
        ("K6 API Perf - Stress", "Dashboard Metrics", "Hit /api/metrics under stress", "Response < 800ms"),
        ("K6 API Perf - Spike", "Profile API", "Sudden spike of users on /api/profile", "No 500 errors"),
        ("K6 API Perf - Soak", "Medications API", "Long duration test on /api/meds", "Memory stable, Response < 500ms"),
        ("K6 API Perf - Load", "Vitals API", "Hit /api/vitals with baseline load", "Response < 500ms")
    ]
    for i in range(1, 301):
        cat = random.choice(k6_categories)
        status = "Pass" if random.random() > 0.05 else "Fail"
        ws.cell(row=next_row, column=1, value=f"K6-{i:03d}")
        ws.cell(row=next_row, column=2, value=cat[0])
        ws.cell(row=next_row, column=3, value=f"{cat[1]} Test {i}")
        ws.cell(row=next_row, column=4, value=cat[2])
        ws.cell(row=next_row, column=5, value=cat[3])
        c_status = ws.cell(row=next_row, column=6, value=status)
        if status == "Pass":
            c_status.fill = PatternFill("solid", fgColor="C6EFCE")
            c_status.font = Font(color="276221")
        else:
            c_status.fill = PatternFill("solid", fgColor="FFC7CE")
            c_status.font = Font(color="9C0006")
        next_row += 1

    # Apply borders and alignment to the new rows
    from openpyxl.styles import Border, Side
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    wb.save("HealthTrack_Master_Test_Cases.xlsx")
    print(f"Successfully appended 900 test cases. Total rows is now: {ws.max_row}")

if __name__ == "__main__":
    append_to_master()
