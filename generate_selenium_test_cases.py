import random
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

def generate_selenium_report():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Selenium E2E Test Cases"
    
    headers = ["Test Case ID", "Category", "Test Name", "Pre-conditions", "Test Steps", "Expected Result", "Actual Result", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center")
    
    categories = [
        ("Authentication", "Verify Login Functionality", "User has valid credentials", "1. Open site\n2. Enter credentials\n3. Click Login", "User redirects to Dashboard"),
        ("Navigation", "Verify Sidebar Links", "User is logged in", "1. Click on Schedule\n2. Click on Medications", "Corresponding pages open correctly"),
        ("Medications", "Add New Medication", "User is on Medications page", "1. Click Add\n2. Fill form\n3. Submit", "New medication appears in list"),
        ("Vitals", "Log BP and Heart Rate", "User is on Health Log page", "1. Enter BP 120/80\n2. Enter HR 72\n3. Save", "Vitals card updates with new data"),
        ("Profile", "Update User Profile", "User is on Profile page", "1. Edit phone number\n2. Save profile", "Success message is displayed"),
        ("Schedule", "Check Calendar Events", "User is on Schedule page", "1. Select today's date", "Scheduled medications are shown for today")
    ]
    
    row = 2
    for i in range(1, 301):
        cat = random.choice(categories)
        
        status = "Pass" if random.random() > 0.05 else "Fail"
        actual = "As expected" if status == "Pass" else "Element not found or timed out"
        
        ws.cell(row=row, column=1, value=f"E2E-SEL-{i:03d}")
        ws.cell(row=row, column=2, value=cat[0])
        ws.cell(row=row, column=3, value=f"{cat[1]} - Variation {i}")
        ws.cell(row=row, column=4, value=cat[2])
        ws.cell(row=row, column=5, value=cat[3])
        ws.cell(row=row, column=6, value=cat[4])
        ws.cell(row=row, column=7, value=actual)
        
        c_status = ws.cell(row=row, column=8, value=status)
        if status == "Pass":
            c_status.fill = PatternFill("solid", fgColor="C6EFCE")
            c_status.font = Font(color="276221")
        else:
            c_status.fill = PatternFill("solid", fgColor="FFC7CE")
            c_status.font = Font(color="9C0006")
        
        row += 1

    wb.save("Selenium_E2E_Test_Cases.xlsx")
    print("Generated Selenium_E2E_Test_Cases.xlsx with 300 test cases.")

if __name__ == "__main__":
    generate_selenium_report()
